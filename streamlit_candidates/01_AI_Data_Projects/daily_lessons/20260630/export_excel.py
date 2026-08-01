import streamlit as st

import os
import sys
import re
from html.parser import HTMLParser
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class MovieParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.movies = []
        self.current_movie = None
        self.in_h2 = False
        self.in_category = False
        self.in_info = False
        self.in_score = False
        
        self.current_categories = []
        self.current_info_spans = []

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        
        if tag == "div" and "item" in attrs_dict.get("class", ""):
            self.save_current_movie()
            self.current_movie = {}
            self.current_categories = []
            self.current_info_spans = []
            
        if self.current_movie is None:
            return
            
        if tag == "img" and attrs_dict.get("class") == "cover":
            self.current_movie["cover"] = attrs_dict.get("src", "")
            
        elif tag == "h2" and attrs_dict.get("class") == "m-b-sm":
            self.in_h2 = True
            
        elif tag == "button" and "category" in attrs_dict.get("class", ""):
            self.in_category = True
            
        elif tag == "div" and "info" in attrs_dict.get("class", ""):
            self.in_info = True
            
        elif tag == "p" and "score" in attrs_dict.get("class", ""):
            self.in_score = True

    def handle_data(self, data):
        if self.current_movie is None:
            return
            
        data_str = data.strip()
        if not data_str:
            return
            
        if self.in_h2:
            self.current_movie["raw_title"] = data_str
            if " - " in data_str:
                parts = data_str.split(" - ", 1)
                self.current_movie["title_zh"] = parts[0].strip()
                self.current_movie["title_en"] = parts[1].strip()
            else:
                self.current_movie["title_zh"] = data_str
                self.current_movie["title_en"] = ""
                
        elif self.in_category:
            self.current_categories.append(data_str)
            
        elif self.in_info:
            if data_str != "/":
                self.current_info_spans.append(data_str)
                
        elif self.in_score:
            self.current_movie["score"] = data_str

    def handle_endtag(self, tag):
        if self.current_movie is None:
            return
            
        if tag == "h2" and self.in_h2:
            self.in_h2 = False
        elif tag == "button" and self.in_category:
            self.in_category = False
        elif tag == "div" and self.in_info:
            self.in_info = False
        elif tag == "p" and self.in_score:
            self.in_score = False

    def save_current_movie(self):
        if self.current_movie:
            self.current_movie["categories"] = "/".join(self.current_categories)
            
            region = ""
            duration = ""
            release_date = ""
            
            info_items = [item for item in self.current_info_spans if item and item != "/"]
            
            if len(info_items) >= 1:
                region = info_items[0]
            if len(info_items) >= 2:
                duration = info_items[1]
            if len(info_items) >= 3:
                release_date = info_items[2]
                
            self.current_movie["region"] = region
            self.current_movie["duration"] = duration
            self.current_movie["release_date"] = release_date
            
            self.movies.append(self.current_movie)
            self.current_movie = None

    def close(self):
        self.save_current_movie()
        super().close()

def sanitize_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

def export_to_excel():
    html_path = "page1.html"
    if not os.path.exists(html_path):
        st.write(f"Error: '{html_path}' does not exist. Please run crawler.py first.")
        return
        
    with open(html_path, "r", encoding="utf-8") as f:
        html_content = f.read()

    parser = MovieParser()
    parser.feed(html_content)
    parser.close()
    
    movies = parser.movies

    if not movies:
        st.write("No movies found to export!")
        return

    excel_path = "movies.xlsx"
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scraped Movies"
    
    # Show grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    cell_font = Font(name=font_family, size=10)
    score_font = Font(name=font_family, size=11, bold=True)
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    even_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    headers = [
        "Index",
        "Poster",
        "Chinese Title",
        "English Title",
        "Score",
        "Categories",
        "Region",
        "Duration",
        "Release Date"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # Set header row height
    ws.row_dimensions[1].height = 28
    
    # Set column widths (characters)
    column_widths = {
        1: 8,   # Index
        2: 15,  # Poster Image
        3: 25,  # Chinese Title
        4: 30,  # English Title
        5: 8,   # Score
        6: 22,  # Categories
        7: 25,  # Region
        8: 12,  # Duration
        9: 18   # Release Date
    }
    for col_num, width in column_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width

    st.write(f"Exporting movies to Excel with embedded posters...")
    
    for idx, movie in enumerate(movies, 1):
        row_num = idx + 1
        
        # Determine cell values
        title_zh = movie.get("title_zh", "")
        title_en = movie.get("title_en", "")
        score = float(movie.get("score", "0")) if movie.get("score") else ""
        categories = movie.get("categories", "")
        region = movie.get("region", "")
        duration = movie.get("duration", "")
        release_date = movie.get("release_date", "")
        cover_url = movie.get("cover", "")
        
        # Write values
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=3, value=title_zh)
        ws.cell(row=row_num, column=4, value=title_en)
        ws.cell(row=row_num, column=5, value=score)
        ws.cell(row=row_num, column=6, value=categories)
        ws.cell(row=row_num, column=7, value=region)
        ws.cell(row=row_num, column=8, value=duration)
        ws.cell(row=row_num, column=9, value=release_date)
        
        # Styles for row cells
        fill = even_row_fill if row_num % 2 == 0 else odd_row_fill
        ws.row_dimensions[row_num].height = 105 # Set row height in points (approx 140 pixels)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = score_font if col_num == 5 else cell_font
            cell.fill = fill
            cell.border = thin_border
            
            # Alignments
            if col_num in [1, 2, 5, 8, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
        # Insert image in Column 2 (Poster)
        safe_title = sanitize_filename(title_zh)
        ext = ".jpg"
        if ".png" in cover_url.lower():
            ext = ".png"
        local_path = f"posters/{idx}_{safe_title}{ext}"
        
        if os.path.exists(local_path):
            try:
                # Openpyxl drawing Image
                img = OpenpyxlImage(local_path)
                # Scale image to height=120px, width=86px (maintains aspect ratio ~1.39)
                img.height = 120
                img.width = 86
                # Add to cell B{row}
                ws.add_image(img, f"B{row_num}")
            except Exception as e:
                st.write(f"  Error inserting image for row {idx}: {e}")
                ws.cell(row=row_num, column=2, value="Error Loading Image")
        else:
            ws.cell(row=row_num, column=2, value="No Image Downloaded")

    try:
        wb.save(excel_path)
        st.write(f"Success! Generated Excel spreadsheet: {excel_path}")
    except Exception as e:
        st.write(f"Failed to save Excel file: {e}", file=sys.stderr)

if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    export_to_excel()
