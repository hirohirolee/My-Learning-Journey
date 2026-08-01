import streamlit as st

import os
import csv
import sys
import re
import ssl
import urllib.request
import urllib.error
from html.parser import HTMLParser
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Define standard headers to mimic a browser
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# Create custom SSL context to bypass expired/invalid SSL certificates
ssl_context = ssl.create_default_context()
ssl_context.check_hostname = False
ssl_context.verify_mode = ssl.CERT_NONE

class MovieParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.movies = []
        self.current_movie = None
        self.in_h2 = False
        self.in_category = False
        self.in_info = False
        self.in_score = False
        
        self.current_categories = []
        self.current_info_spans = []

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        
        if tag == "div" and "item" in attrs_dict.get("class", ""):
            self.save_current_movie()
            self.current_movie = {}
            self.current_categories = []
            self.current_info_spans = []
            
        if self.current_movie is None:
            return
            
        if tag == "img" and attrs_dict.get("class") == "cover":
            self.current_movie["cover"] = attrs_dict.get("src", "")
            
        elif tag == "h2" and attrs_dict.get("class") == "m-b-sm":
            self.in_h2 = True
            
        elif tag == "button" and "category" in attrs_dict.get("class", ""):
            self.in_category = True
            
        elif tag == "div" and "info" in attrs_dict.get("class", ""):
            self.in_info = True
            
        elif tag == "p" and "score" in attrs_dict.get("class", ""):
            self.in_score = True

    def handle_data(self, data):
        if self.current_movie is None:
            return
            
        data_str = data.strip()
        if not data_str:
            return
            
        if self.in_h2:
            self.current_movie["raw_title"] = data_str
            if " - " in data_str:
                parts = data_str.split(" - ", 1)
                self.current_movie["title_zh"] = parts[0].strip()
                self.current_movie["title_en"] = parts[1].strip()
            else:
                self.current_movie["title_zh"] = data_str
                self.current_movie["title_en"] = ""
                
        elif self.in_category:
            self.current_categories.append(data_str)
            
        elif self.in_info:
            if data_str != "/":
                self.current_info_spans.append(data_str)
                
        elif self.in_score:
            self.current_movie["score"] = data_str

    def handle_endtag(self, tag):
        if self.current_movie is None:
            return
            
        if tag == "h2" and self.in_h2:
            self.in_h2 = False
        elif tag == "button" and self.in_category:
            self.in_category = False
        elif tag == "div" and self.in_info:
            self.in_info = False
        elif tag == "p" and self.in_score:
            self.in_score = False

    def save_current_movie(self):
        if self.current_movie:
            self.current_movie["categories"] = "/".join(self.current_categories)
            
            region = ""
            duration = ""
            release_date = ""
            
            info_items = [item for item in self.current_info_spans if item and item != "/"]
            
            if len(info_items) >= 1:
                region = info_items[0]
            if len(info_items) >= 2:
                duration = info_items[1]
            if len(info_items) >= 3:
                release_date = info_items[2]
                
            self.current_movie["region"] = region
            self.current_movie["duration"] = duration
            self.current_movie["release_date"] = release_date
            
            self.movies.append(self.current_movie)
            self.current_movie = None

    def close(self):
        self.save_current_movie()
        super().close()

def sanitize_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

def download_url_content(url):
    """Downloads content from a URL bypassing SSL validation."""
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, context=ssl_context) as response:
        return response.read()

def crawl_pages(max_pages=10):
    cache_dir = "cache"
    os.makedirs(cache_dir, exist_ok=True)
    all_movies = []
    
    st.write(f"--- Phase 1: Scraping HTML Pages (1 to {max_pages}) ---")
    for page in range(1, max_pages + 1):
        cache_path = os.path.join(cache_dir, f"page_{page}.html")
        url = f"https://ssr1.scrape.center/page/{page}"
        
        if os.path.exists(cache_path):
            st.write(f"Page {page} found in cache. Loading...")
            with open(cache_path, "r", encoding="utf-8") as f:
                html_content = f.read()
        else:
            st.write(f"Downloading Page {page} from {url}...")
            try:
                content = download_url_content(url)
                html_content = content.decode("utf-8")
                with open(cache_path, "w", encoding="utf-8") as f:
                    f.write(html_content)
            except Exception as e:
                st.write(f"Error downloading Page {page}: {e}", file=sys.stderr)
                continue
                
        parser = MovieParser()
        parser.feed(html_content)
        parser.close()
        st.write(f"  Parsed {len(parser.movies)} movies from Page {page}.")
        all_movies.extend(parser.movies)
        
    return all_movies[:100] # Ensure we keep exactly 100 movies if there are more

def download_single_poster(idx, movie, posters_dir):
    title = movie.get("title_zh", "")
    img_url = movie.get("cover", "")
    
    if not img_url:
        return idx, title, False, "No URL"
        
    safe_title = sanitize_filename(title)
    ext = ".jpg"
    if ".png" in img_url.lower():
        ext = ".png"
    filename = f"{idx}_{safe_title}{ext}"
    filepath = os.path.join(posters_dir, filename)
    
    if os.path.exists(filepath):
        return idx, title, True, "Already exists"
        
    try:
        data = download_url_content(img_url)
        with open(filepath, "wb") as f:
            f.write(data)
        return idx, title, True, filepath
    except Exception as e:
        return idx, title, False, str(e)

def download_posters_concurrently(movies, posters_dir="posters", max_workers=10):
    os.makedirs(posters_dir, exist_ok=True)
    st.write(f"\n--- Phase 2: Downloading Poster Images (Concurrently) ---")
    
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(download_single_poster, idx, movie, posters_dir): idx 
            for idx, movie in enumerate(movies, 1)
        }
        
        for future in as_completed(futures):
            idx, title, success, msg = future.result()
            if success:
                if msg == "Already exists":
                    pass # silent if cached
                else:
                    st.write(f"  [{idx}/100] Downloaded poster for: {title}")
            else:
                st.write(f"  [{idx}/100] Failed poster download for {title}: {msg}", file=sys.stderr)
            results[idx] = success
            
    st.write("Poster downloads completed.")

def export_to_csv(movies, csv_path="movies.csv"):
    st.write(f"\n--- Phase 3: Exporting to CSV ---")
    headers = [
        "Index",
        "Poster (Excel Formula)",
        "Chinese Title",
        "English Title",
        "Score",
        "Categories",
        "Region",
        "Duration",
        "Release Date",
        "Poster URL",
        "Local Poster Path"
    ]
    
    try:
        f = open(csv_path, "w", newline="", encoding="utf-8-sig")
    except Exception as e:
        st.write(f"Failed to open primary CSV file {csv_path} (it may be open/locked): {e}", file=sys.stderr)
        f = None
        for i in range(1, 10):
            fallback_path = f"movies_{i}.csv"
            try:
                f = open(fallback_path, "w", newline="", encoding="utf-8-sig")
                csv_path = fallback_path
                st.write(f"Will save CSV to fallback file: {csv_path}")
                break
            except Exception:
                continue
        if not f:
            st.write("Failed to save CSV file to any location.", file=sys.stderr)
            return

    try:
        with f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            for idx, movie in enumerate(movies, 1):
                title_zh = movie.get("title_zh", "")
                title_en = movie.get("title_en", "")
                score = movie.get("score", "")
                categories = movie.get("categories", "")
                region = movie.get("region", "")
                duration = movie.get("duration", "")
                release_date = movie.get("release_date", "")
                cover_url = movie.get("cover", "")
                
                safe_title = sanitize_filename(title_zh)
                ext = ".jpg"
                if ".png" in cover_url.lower():
                    ext = ".png"
                local_path = f"posters/{idx}_{safe_title}{ext}"
                if not os.path.exists(local_path):
                    local_path = ""
                    
                image_formula = f'=IMAGE("{cover_url}")'
                
                writer.writerow([
                    idx,
                    image_formula,
                    title_zh,
                    title_en,
                    score,
                    categories,
                    region,
                    duration,
                    release_date,
                    cover_url,
                    local_path
                ])
        st.write(f"Success! Saved 100 movies to CSV: {csv_path}")
    except Exception as e:
        st.write(f"Failed to export CSV: {e}", file=sys.stderr)

def export_to_excel(movies, excel_path="movies.xlsx"):
    st.write(f"\n--- Phase 4: Exporting to Excel with Embedded Images ---")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scraped Movies (100)"
    
    # Show grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Fonts and fills
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    cell_font = Font(name=font_family, size=10)
    score_font = Font(name=font_family, size=11, bold=True)
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    even_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    headers = [
        "Index",
        "Poster",
        "Chinese Title",
        "English Title",
        "Score",
        "Categories",
        "Region",
        "Duration",
        "Release Date"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    ws.row_dimensions[1].height = 28
    
    column_widths = {
        1: 8,   # Index
        2: 15,  # Poster Image
        3: 25,  # Chinese Title
        4: 30,  # English Title
        5: 8,   # Score
        6: 22,  # Categories
        7: 25,  # Region
        8: 12,  # Duration
        9: 18   # Release Date
    }
    for col_num, width in column_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width

    for idx, movie in enumerate(movies, 1):
        row_num = idx + 1
        
        title_zh = movie.get("title_zh", "")
        title_en = movie.get("title_en", "")
        score = float(movie.get("score", "0")) if movie.get("score") else ""
        categories = movie.get("categories", "")
        region = movie.get("region", "")
        duration = movie.get("duration", "")
        release_date = movie.get("release_date", "")
        cover_url = movie.get("cover", "")
        
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=3, value=title_zh)
        ws.cell(row=row_num, column=4, value=title_en)
        ws.cell(row=row_num, column=5, value=score)
        ws.cell(row=row_num, column=6, value=categories)
        ws.cell(row=row_num, column=7, value=region)
        ws.cell(row=row_num, column=8, value=duration)
        ws.cell(row=row_num, column=9, value=release_date)
        
        fill = even_row_fill if row_num % 2 == 0 else odd_row_fill
        ws.row_dimensions[row_num].height = 105
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = score_font if col_num == 5 else cell_font
            cell.fill = fill
            cell.border = thin_border
            
            if col_num in [1, 2, 5, 8, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
        # Insert image
        safe_title = sanitize_filename(title_zh)
        ext = ".jpg"
        if ".png" in cover_url.lower():
            ext = ".png"
        local_path = f"posters/{idx}_{safe_title}{ext}"
        
        if os.path.exists(local_path):
            try:
                img = OpenpyxlImage(local_path)
                img.height = 120
                img.width = 86
                ws.add_image(img, f"B{row_num}")
            except Exception as e:
                ws.cell(row=row_num, column=2, value="Error Loading")
        else:
            ws.cell(row=row_num, column=2, value="No Poster")
            
    try:
        wb.save(excel_path)
        st.write(f"Success! Saved 100 movies with embedded posters to Excel: {excel_path}")
    except Exception as e:
        st.write(f"Failed to save Excel file to {excel_path} (it may be open/locked): {e}", file=sys.stderr)
        # Try fallback names
        for i in range(1, 10):
            fallback_path = f"movies_{i}.xlsx"
            try:
                wb.save(fallback_path)
                st.write(f"Success! Saved to fallback file: {fallback_path}")
                break
            except Exception as fe:
                continue

def main():
    # Fix encoding
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
        
    movies = crawl_pages(max_pages=10)
    st.write(f"\nSuccessfully parsed a total of {len(movies)} movies.")
    
    if len(movies) > 0:
        download_posters_concurrently(movies, posters_dir="posters", max_workers=15)
        export_to_csv(movies, csv_path="movies.csv")
        export_to_excel(movies, excel_path="movies.xlsx")
        st.write("\nAll tasks completed successfully!")

if __name__ == "__main__":
    main()
